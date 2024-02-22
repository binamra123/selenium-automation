from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import pytest
import openpyxl
import time

def readData():
    data_list = []
    path = "C:\\Users\\Admin\\OneDrive\\Desktop\\testlogin.xlsx" 

    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']

    rows = sheet.max_row

    for r in range(2, rows+1):
        username = sheet.cell(row=r, column=1).value
        password = sheet.cell(row=r, column=2).value
        data_list.append((username, password))

    return data_list

@pytest.fixture
def setup():
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.implicitly_wait(10)
    yield driver
    driver.quit()

@pytest.mark.parametrize("username, password", readData())
def test_login(setup, username, password):
    driver = setup
    driver.get("https://practice.expandtesting.com/login")
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)  
    driver.find_element(By.XPATH, "//button[@class='btn btn-bg btn-primary d-block w-100']").click()
    time.sleep(3)

def test_complete():
    print('Test Completed')
