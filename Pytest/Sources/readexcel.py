import openpyxl

def readData():
    list = []
    path = "C:\\Users\\Admin\\OneDrive\\Desktop\\testlogin.xlsx"

    workbook = openpyxl.load_workbook(path)

    sheet = workbook.get_sheet_by_name('Sheet1')

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows+1):
         username = sheet.cell(r, 1).value
         password = sheet.cell(r, 2).value


    tuple = (username, password) 
    list.append(tuple)
 
print(list)
 


   
